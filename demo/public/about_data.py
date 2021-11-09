'''
Author: tyler
Date: 2021-08-18 16:08:10
LastEditTime: 2021-11-01 15:17:52
LastEditors: Please set LastEditors
Description: Data of read and save
FilePath: \tylerhub\demo\public\about_data.py
'''
import os
import sys

import xlrd
import xlwt
from openpyxl import load_workbook
from collections import Counter

class Aboutdata(object):

    """
    此模块用于提取excel表格中的数据，并封装成含多个字典的列表
    文件格式为xlsx的后缀即可
    """

    def openexcel(self,excelpath,sheetname):
        """excelpath为excel文件存储路径，sheetname为sheet名"""
        #打开excel表格
        self.data=xlrd.open_workbook(excelpath)
        #通过sheet名打开对应sheet表
        self.table=self.data.sheet_by_name(sheetname)
        #获取第一行数据作为key值,下标为0
        self.key=self.table.row_values(0)
        #获取总行数
        self.rows=self.table.nrows
        #获取总列数
        self.clos=self.table.ncols
        #返回行数
        return self.rows


    #将表中数据以含多个字典的列表的数据类型输出
    def dict_data(self):
        #判断表中有无数据
        if self.rows <= 1:
            print('该sheet表无数据')
        else:
            l=[]
            j=1
            for i in range(self.rows-1): #去除首行
                self.values=self.table.row_values(j)#第j行数据作为values
                s={}
                for x in range(self.clos):#根据列数
                    s[self.key[x]]=self.values[x] #每个循环的字典中，key列表中的第x项=valus列表中的第x项
                l.append(s) #i循环下，添加进列表l
                j=j+1
            return l


    # 创建存储注册数据的函数，写入已存在的本地文档中,cloumn:列；row:行
    def saveainfo(self,excelpath,values,column,row:int)->str:
        """
        param:column 列
        param:row 行
        注：调用此函数时，不能打开需要写入数据的文档
        """
        try:
            workbook = load_workbook(filename=excelpath)
            sheet = workbook.active
            cel = sheet['{}{}'.format(column, row)]
            cel.value = values
            workbook.save(filename=excelpath)
        except Exception as msg:
            print('请勿打开需要写入文件的文档：{}'.format(msg))


    def str_insert(self,old_str,pos:int,add_str)->str:
        """
        字符串指定位置添加字符串
        :param old_str : 被添加字符的字符串
        :param pos : 添加的位置
        :param add_str : 添加的字符串
        :return 返回新字符串
        """
        self.str_list=list(old_str)
        self.str_list.insert(pos, add_str)
        self.str_out=''.join(self.str_list)
        return self.str_out

    def txt_xlsx(self,filename,xlsxname):   
        """txt转换成xlsx 
        :param filename txt文本文件名称、  
        :param xlsxname 表示转换后的excel文件名  
        """   
        try:      
            f = open(filename,'r')
            f.seek(0,0) #游标回到txt文本头部
            xls=xlwt.Workbook()        
            #生成excel的方法，声明excel        
            sheet = xls.add_sheet('sheet1',cell_overwrite_ok=True)        
            x = 0
            while True:            
                #按行循环，读取文本文件           
                line = f.readline()            
                if not line:               
                    break  #如果没有内容，则退出循环            
                for i in range(len(line.split('\t'))):             
                    item=line.split('\t')[i]               
                    sheet.write(x,i,item)
                    #x单元格经度，i 单元格纬度           
                x += 1 #excel另起一行      
            f.close()        
            xls.save(xlsxname)
            #保存xls文件   
        except:        
            raise

#测试
if __name__=='__main__':
    dealData=Aboutdata()
    # e.openexcel(r'C:\Users\tyler.tang\Desktop\code\tylerhub\demo\cl_open_demoaccount\test_data\cl_open_demo.xlsx','Sheet1')
    # a=e.dict_data()
    # print(a)
    # print(a[0]['主账号'])
    #e.saveainfo(r'C:\Users\tyler.tang\Desktop\code\tylerhub\demo\registration_process\test_excel_data\Account_number.xlsx', 'd', 'A', 4)
    # dealData.openexcel(r'C:\Users\tyler.tang\Desktop\code\tylerhub\demo\order_list\test_data\copy_openOrder.xlsx','sheet1')
    # data=dealData.dict_data()
    # print(data)
    # import pyperclip
    # data=pyperclip.paste()
    # li=[]
    # for i in data:
    #     li.append(i.strip('\n'))

    # with open(r'C:\Users\tyler.tang\Desktop\code\tylerhub\demo\order_list\test_data\test.txt', 'w+') as f:
    #     f.write(''.join(li))

    # filename=r'C:\Users\tyler.tang\Desktop\code\tylerhub\demo\order_list\test_data\test.txt'
    # xlsxname=r'C:\Users\tyler.tang\Desktop\code\tylerhub\demo\order_list\test_data\Open_orderlist.xlsx'
    # dealData.txt_xlsx(filename,xlsxname)
    dealData.openexcel(r'C:\Users\tyler.tang\Desktop\code\tylerhub\demo\order_list\test_data\Open_orderlist.xlsx', 'sheet1')
    data=dealData.dict_data()
    print(data)
